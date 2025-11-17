import os
from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from DamianApp.forms import MarcaForm, ModeloForm, AutoForm, VendedorForm
from DamianApp.models import Marca, Modelo, Auto, Vendedor
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment
from django.db.models import Q 



@login_required
def inicio(request):
    return render(request, 'index.html')


@login_required
def marcas(request):
    search_query = request.GET.get('q', '') 
    
    if search_query:
        marcas_list = Marca.objects.filter(nombre__icontains=search_query).order_by('nombre')
    else:
        marcas_list = Marca.objects.all().order_by('nombre')
    
    paginator = Paginator(marcas_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    data = {
        'titulo': 'Lista de Marcas',
        'page_obj': page_obj,
        'search_query': search_query 
    }
    return render(request, 'Vehiculos/marcas.html', data)

@login_required
def modelos(request):
    search_query = request.GET.get('q', '')
    
    if search_query:
        modelos_list = Modelo.objects.filter(
            Q(nombre__icontains=search_query) |
            Q(tipo__icontains=search_query) |
            Q(version__icontains=search_query)
        ).order_by('nombre')
    else:
        modelos_list = Modelo.objects.all().order_by('nombre')
    
    paginator = Paginator(modelos_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    data = {
        'titulo': 'Gestión de Modelos',
        'page_obj': page_obj,
        'search_query': search_query
    }
    return render(request, 'Vehiculos/modelos.html', data)

@login_required
def autos(request):
    search_query = request.GET.get('q', '')

    if search_query:
        autos_list = Auto.objects.filter(
            Q(patente__icontains=search_query) |
            Q(estado__icontains=search_query) |
            Q(modelo__nombre__icontains=search_query) |
            Q(marca__nombre__icontains=search_query)
        )
    else:
        autos_list = Auto.objects.all()
    
    paginator = Paginator(autos_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    data = {
        'titulo': 'Gestión de Autos',
        'page_obj': page_obj,
        'search_query': search_query
    }
    return render(request, 'Vehiculos/autos.html', data)

@login_required
def vendedores(request):
    search_query = request.GET.get('q', '')
    
    if search_query:
        vendedores_list = Vendedor.objects.filter(
            Q(run__icontains=search_query) |
            Q(nombre__icontains=search_query)
        ).order_by('nombre')
    else:
        vendedores_list = Vendedor.objects.all().order_by('nombre')
    
    paginator = Paginator(vendedores_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    data = {
        'titulo': 'Lista de Vendedores',
        'page_obj': page_obj,
        'search_query': search_query
    }
    return render(request, 'Vehiculos/vendedores.html', data)

@permission_required('DamianApp.add_marca', login_url='accounts/login')
def crearMarca(request):
    form = MarcaForm()
    data = {'titulo': 'Crear Marca', 'form': form, 'ruta': '/Vehiculos/marcas/'}
    if request.method == 'POST':
        form = MarcaForm(request.POST, request.FILES)
        if form.is_valid():
            FileSystemStorage(location='media/marcas')
            form.save()
            messages.success(request, 'Marca creada con éxito!')
    return render(request, 'Vehiculos/createF.html', data)

@permission_required('DamianApp.change_marca', login_url='accounts/login')
def editarMarca(request, id):
    marca = Marca.objects.get(pk=id)
    old_logo_path = None
    if marca.logo:
        old_logo_path = marca.logo.path
    form = MarcaForm(instance=marca)
    data = {'titulo': 'Editar Marca', 'form': form, 'ruta': '/Vehiculos/marcas/'}
    if request.method == 'POST':
        form = MarcaForm(request.POST, request.FILES, instance=marca)
        if form.is_valid():
            new_logo = request.FILES.get('logo') 
            form.save() 
            messages.success(request, 'Marca editada con éxito!')
            if new_logo and old_logo_path:
                if marca.logo.path != old_logo_path:
                    if os.path.exists(old_logo_path):
                        os.remove(old_logo_path)
    return render(request, 'Vehiculos/createF.html', data)

@permission_required('DamianApp.delete_marca', login_url='accounts/login')
def eliminarMarca(request, id):
    marca = Marca.objects.get(pk=id)
    marca.delete()
    return redirect('/Vehiculos/marcas')


@permission_required('DamianApp.add_modelo', login_url='accounts/login')
def crearModelo(request):
    form = ModeloForm()
    data = {'titulo': 'Crear Modelo', 'form': form, 'ruta': '/Vehiculos/modelos/'}
    if request.method == 'POST':
        form = ModeloForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modelo creado con éxito!')
    return render(request, 'Vehiculos/create.html', data)

@permission_required('DamianApp.change_modelo', login_url='accounts/login')
def editarModelo(request, id):
    modelo = Modelo.objects.get(pk=id)
    form = ModeloForm(instance=modelo)
    data = {'titulo': 'Editar Modelo', 'form': form, 'ruta': '/Vehiculos/modelos/'}
    if request.method == 'POST':
        form = ModeloForm(request.POST, instance=modelo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Modelo editado con éxito!')
    return render(request, 'Vehiculos/create.html', data)

@permission_required('DamianApp.delete_modelo', login_url='accounts/login')
def eliminarModelo(request, id):
    modelo = Modelo.objects.get(pk=id)
    modelo.delete()
    return redirect('/Vehiculos/modelos')


@permission_required('DamianApp.add_auto', login_url='accounts/login')
def crearAuto(request):
    form = AutoForm()
    data = {'titulo': 'Crear Auto', 'form': form, 'ruta': '/Vehiculos/autos/'}
    if request.method == 'POST' and request.FILES.get('foto'):
        form = AutoForm(request.POST, request.FILES)
        if form.is_valid():
            FileSystemStorage(location='media/autos')
            form.save()
            messages.success(request, 'Auto creado con éxito!')
        else:
            messages.error(request, 'Error al crear el auto')
    return render(request, 'Vehiculos/createF.html', data)

@permission_required('DamianApp.change_auto', login_url='accounts/login')
def editarAuto(request, patente):
    auto = Auto.objects.get(patente=patente)
    old_foto_path = None
    if auto.foto:
        old_foto_path = auto.foto.path
    form = AutoForm(instance=auto)
    data = {'titulo': 'Editar Auto', 'form': form, 'ruta': '/Vehiculos/autos/'}
    if request.method == 'POST':
        form = AutoForm(request.POST, request.FILES, instance=auto)
        if form.is_valid():
            new_foto = request.FILES.get('foto')
            form.save() 
            messages.success(request, 'Auto editado con éxito!')
            if new_foto and old_foto_path:
                if auto.foto.path != old_foto_path:
                    if os.path.exists(old_foto_path):
                        os.remove(old_foto_path)
    return render(request, 'Vehiculos/createF.html', data)

@permission_required('DamianApp.delete_auto', login_url='accounts/login')
def eliminarAuto(request, patente):
    auto = Auto.objects.get(patente=patente)
    auto.delete()
    return redirect('/Vehiculos/autos')


@permission_required('DamianApp.add_vendedor', login_url='accounts/login')
def crearVendedor(request):
    form = VendedorForm()
    data = {'titulo': 'Crear Vendedor', 'form': form, 'ruta': '/Vehiculos/vendedores/'}
    if request.method == 'POST':
        form = VendedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vendedor creado con éxito!')
    return render(request, 'Vehiculos/create.html', data)

@permission_required('DamianApp.change_vendedor', login_url='accounts/login')
def editarVendedor(request, run):
    vendedor = Vendedor.objects.get(run=run)
    form = VendedorForm(instance=vendedor)
    data = {'titulo': 'Editar Vendedor', 'form': form, 'ruta': '/Vehiculos/vendedores/'}
    if request.method == 'POST':
        form = VendedorForm(request.POST, instance=vendedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vendedor editado con éxito!')
    return render(request, 'Vehiculos/create.html', data)

@permission_required('DamianApp.delete_vendedor', login_url='accounts/login')
def eliminarVendedor(request, run):
    vendedor = Vendedor.objects.get(run=run)
    vendedor.delete()
    return redirect('/Vehiculos/vendedores')


def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save() 
            login(request, user) 
            messages.success(request, '¡Registro exitoso! Has iniciado sesión.')
            return redirect('inicio')
        else:
            messages.error(request, 'Por favor corrige los errores para registrarte.')
    else:
        form = UserCreationForm()
    data = {'form': form, 'titulo': 'Registro de Usuario'}
    return render(request, 'registration/register.html', data)



@permission_required('DamianApp.view_marca', login_url='accounts/login')
def export_marcas_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_marcas.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marcas"
    header_font = Font(bold=True, size=12)
    ws.cell(row=1, column=1, value="ID").font = header_font
    ws.cell(row=1, column=2, value="Nombre").font = header_font
    ws.cell(row=1, column=3, value="Descripción").font = header_font
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60
    marcas = Marca.objects.all().order_by('id')
    cont = 2 
    for m in marcas:
        ws.cell(row=cont, column=1, value=m.id)
        ws.cell(row=cont, column=2, value=m.nombre)
        ws.cell(row=cont, column=3, value=m.descripcion)
        cont += 1
    wb.save(response)
    return response

@permission_required('DamianApp.view_modelo', login_url='accounts/login')
def export_modelos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_modelos.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelos"
    header_font = Font(bold=True, size=12)
    ws.cell(row=1, column=1, value="ID").font = header_font
    ws.cell(row=1, column=2, value="Nombre").font = header_font
    ws.cell(row=1, column=3, value="Tipo").font = header_font
    ws.cell(row=1, column=4, value="Versión").font = header_font
    ws.cell(row=1, column=5, value="Tipo de Motor").font = header_font
    ws.cell(row=1, column=6, value="Años Fabricación").font = header_font
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    modelos = Modelo.objects.all().order_by('id')
    cont = 2
    for m in modelos:
        ws.cell(row=cont, column=1, value=m.id)
        ws.cell(row=cont, column=2, value=m.nombre)
        ws.cell(row=cont, column=3, value=m.tipo)
        ws.cell(row=cont, column=4, value=m.version)
        ws.cell(row=cont, column=5, value=m.tipo_motor)
        ws.cell(row=cont, column=6, value=m.años_fabricacion)
        cont += 1
    wb.save(response)
    return response

@permission_required('DamianApp.view_auto', login_url='accounts/login')
def export_autos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_autos.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Autos"
    header_font = Font(bold=True, size=12)
    ws.cell(row=1, column=1, value="Patente").font = header_font
    ws.cell(row=1, column=2, value="Modelo").font = header_font
    ws.cell(row=1, column=3, value="Marca").font = header_font
    ws.cell(row=1, column=4, value="Kilometraje").font = header_font
    ws.cell(row=1, column=5, value="KM por Litro").font = header_font
    ws.cell(row=1, column=6, value="Año").font = header_font
    ws.cell(row=1, column=7, value="Estado").font = header_font
    ws.cell(row=1, column=8, value="Disponibilidad").font = header_font
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 15
    autos = Auto.objects.all().order_by('patente')
    cont = 2
    for a in autos:
        ws.cell(row=cont, column=1, value=a.patente)
        ws.cell(row=cont, column=2, value=str(a.modelo))
        ws.cell(row=cont, column=3, value=str(a.marca))
        ws.cell(row=cont, column=4, value=a.kilometraje)
        ws.cell(row=cont, column=5, value=a.km_por_litro)
        ws.cell(row=cont, column=6, value=a.año)
        ws.cell(row=cont, column=7, value=a.estado)
        ws.cell(row=cont, column=8, value="Sí" if a.disponibilidad else "No")
        cont += 1
    wb.save(response)
    return response

@permission_required('DamianApp.view_vendedor', login_url='accounts/login')
def export_vendedores_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="listado_vendedores.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendedores"
    header_font = Font(bold=True, size=12)
    ws.cell(row=1, column=1, value="RUN").font = header_font
    ws.cell(row=1, column=2, value="Nombre").font = header_font
    ws.cell(row=1, column=3, value="Email").font = header_font
    ws.cell(row=1, column=4, value="Teléfono").font = header_font
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    vendedores = Vendedor.objects.all().order_by('nombre')
    cont = 2
    for v in vendedores:
        ws.cell(row=cont, column=1, value=v.run)
        ws.cell(row=cont, column=2, value=v.nombre)
        ws.cell(row=cont, column=3, value=v.email)
        ws.cell(row=cont, column=4, value=v.telefono)
        cont += 1
    wb.save(response)
    return response